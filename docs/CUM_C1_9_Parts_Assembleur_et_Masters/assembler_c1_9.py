#!/usr/bin/env python3
from __future__ import annotations
import argparse
import copy
import re
from pathlib import Path
from openpyxl import Workbook, load_workbook

INVALID = re.compile(r'[:\\/?*\[\]]')

def unique_sheet_name(name, used, prefix):
    base = INVALID.sub("_", name).strip() or "Feuille"
    candidate = base[:31]
    if candidate not in used:
        used.add(candidate)
        return candidate
    candidate = f"{prefix}_{base}"[:31]
    n = 2
    while candidate in used:
        tail = f"_{n}"
        candidate = (f"{prefix}_{base}")[:31-len(tail)] + tail
        n += 1
    used.add(candidate)
    return candidate

def copy_sheet(src, dst):
    for row in src.iter_rows():
        for c in row:
            d = dst[c.coordinate]
            d.value = c.value
            if c.has_style:
                d._style = copy.copy(c._style)
            if c.hyperlink:
                d._hyperlink = copy.copy(c.hyperlink)
            if c.comment:
                d.comment = copy.copy(c.comment)
    for merged in src.merged_cells.ranges:
        dst.merge_cells(str(merged))
    for key, dim in src.column_dimensions.items():
        dst.column_dimensions[key].width = dim.width
        dst.column_dimensions[key].hidden = dim.hidden
    for idx, dim in src.row_dimensions.items():
        dst.row_dimensions[idx].height = dim.height
        dst.row_dimensions[idx].hidden = dim.hidden
    dst.freeze_panes = src.freeze_panes
    if src.auto_filter and src.auto_filter.ref:
        dst.auto_filter.ref = src.auto_filter.ref

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", default=".")
    ap.add_argument("--output", default="CUM_C1_9_MASTER_LOCAL.xlsx")
    args = ap.parse_args()

    folder = Path(args.input).resolve()
    expected = [
        "CUM_C1_9_1_Faune.xlsx",
        "CUM_C1_9_2_PNJ_Factions.xlsx",
        "CUM_C1_9_3_Rare_Dormant_Obsessions.xlsx",
        "CUM_C1_9_4_Generateur_Integration.xlsx",
    ]
    files = [folder / n for n in expected if (folder / n).exists()]
    missing = [n for n in expected if not (folder / n).exists()]
    if missing:
        raise SystemExit("Fichiers manquants : " + ", ".join(missing))

    out = Workbook()
    out.remove(out.active)
    index = out.create_sheet("00_INDEX_C1_9")
    index.append(["FICHIER_SOURCE","FEUILLE_SOURCE","FEUILLE_MASTER","LIGNES","COLONNES","STATUT"])
    used = {"00_INDEX_C1_9"}

    for fp in files:
        src_wb = load_workbook(fp, data_only=False)
        prefix = re.sub(r"[^A-Za-z0-9]", "", fp.stem)[:10]
        for ws in src_wb.worksheets:
            if ws.title == "00_README" and len(out.sheetnames) > 1:
                continue
            name = unique_sheet_name(ws.title, used, prefix)
            dst = out.create_sheet(name)
            copy_sheet(ws, dst)
            index.append([fp.name, ws.title, name, ws.max_row, ws.max_column, "OK"])
        src_wb.close()

    index.freeze_panes = "A2"
    for col,width in {"A":45,"B":32,"C":32,"D":12,"E":12,"F":18}.items():
        index.column_dimensions[col].width = width
    out.save(folder / args.output)
    print(folder / args.output)

if __name__ == "__main__":
    main()
