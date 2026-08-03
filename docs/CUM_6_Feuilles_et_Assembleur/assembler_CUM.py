from openpyxl import Workbook,load_workbook
from pathlib import Path
src=Path(".")
files=["01_MISSIONS.xlsx","02_OBJECTIFS.xlsx","03_NARRATION.xlsx","04_DEPENDANCES.xlsx","05_MOTEUR.xlsx","06_TABLEAU_DE_BORD.xlsx"]
out=Workbook(); out.remove(out.active)
for f in files:
    wb=load_workbook(src/f)
    ws=wb.active
    nws=out.create_sheet(ws.title)
    for row in ws.iter_rows(values_only=True):
        nws.append(list(row))
out.save("CUM_COMPLET.xlsx")
print("CUM_COMPLET.xlsx créé")
