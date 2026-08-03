#!/usr/bin/env python3
from __future__ import annotations
import argparse, copy, re
from pathlib import Path
from openpyxl import Workbook, load_workbook

INVALID=re.compile(r'[:\\/?*\[\]]')
def unique(name,used,prefix):
    base=INVALID.sub("_",name).strip() or "Feuille"
    cand=base[:31]
    if cand not in used:
        used.add(cand); return cand
    cand=f"{prefix}_{base}"[:31]
    n=2
    while cand in used:
        tail=f"_{n}"; cand=(f"{prefix}_{base}")[:31-len(tail)]+tail; n+=1
    used.add(cand); return cand

def copy_sheet(src,dst):
    for row in src.iter_rows():
        for c in row:
            d=dst[c.coordinate]; d.value=c.value
            if c.has_style: d._style=copy.copy(c._style)
            if c.hyperlink: d._hyperlink=copy.copy(c.hyperlink)
            if c.comment: d.comment=copy.copy(c.comment)
    for m in src.merged_cells.ranges: dst.merge_cells(str(m))
    for k,dim in src.column_dimensions.items():
        dst.column_dimensions[k].width=dim.width; dst.column_dimensions[k].hidden=dim.hidden
    for i,dim in src.row_dimensions.items():
        dst.row_dimensions[i].height=dim.height; dst.row_dimensions[i].hidden=dim.hidden
    dst.freeze_panes=src.freeze_panes

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--input",default=".")
    ap.add_argument("--output",default="CUM_C2_MASTER_LOCAL.xlsx")
    ap.add_argument("--include-global",action="store_true")
    args=ap.parse_args()
    folder=Path(args.input)
    names=[
        "CUM_C2_1_Compilateur.xlsx","CUM_C2_2_Validateur.xlsx","CUM_C2_3_Generateur_Procedural.xlsx",
        "CUM_C2_4_Cartographie_Graphique.xlsx","CUM_C2_5_Documentation_Technique.xlsx"
    ]
    if args.include_global:
        names.insert(0,"CUM_MASTER_C1_9_GLOBAL_ASSEMBLE.xlsx")
    files=[folder/n for n in names if (folder/n).exists()]
    missing=[n for n in names if not (folder/n).exists()]
    if missing:
        raise SystemExit("Fichiers manquants: "+", ".join(missing))
    out=Workbook(); out.remove(out.active)
    idx=out.create_sheet("00_INDEX_C2")
    idx.append(["FICHIER_SOURCE","FEUILLE_SOURCE","FEUILLE_MASTER","LIGNES","COLONNES"])
    used={"00_INDEX_C2"}
    for fp in files:
        wb=load_workbook(fp,data_only=False)
        prefix=re.sub(r"[^A-Za-z0-9]","",fp.stem)[:10]
        for ws in wb.worksheets:
            if ws.title=="00_README" and len(out.sheetnames)>1: continue
            name=unique(ws.title,used,prefix)
            dst=out.create_sheet(name); copy_sheet(ws,dst)
            idx.append([fp.name,ws.title,name,ws.max_row,ws.max_column])
        wb.close()
    out.save(folder/args.output)
    print(folder/args.output)

if __name__=="__main__":
    main()
