#!/usr/bin/env python3
from __future__ import annotations
import argparse, json, hashlib, re
from datetime import datetime, timezone
from pathlib import Path
from openpyxl import load_workbook

def norm(v):
    if v is None:
        return None
    if isinstance(v, str):
        v=v.strip()
        return None if v=="" else v
    return v

def table_rows(ws):
    # Header assumed on row 4 in CUM sheets.
    headers=[norm(c.value) for c in ws[4]]
    if not any(headers):
        return []
    out=[]
    for row in ws.iter_rows(min_row=5, values_only=True):
        obj={}
        empty=True
        for h,v in zip(headers,row):
            if h:
                nv=norm(v)
                obj[h]=nv
                if nv is not None:
                    empty=False
        if not empty:
            out.append(obj)
    return out

DATASETS = {
    "missions":["04_MISSIONS"],
    "mission_children":["04B_MISSIONS_FILLES"],
    "projects":["03_PROJETS","62_PROJETS_AXES"],
    "events":["21_EVENEMENTS"],
    "variables":["31_VARIABLES"],
    "triggers":["42_DECLENCHEURS","73_TRIGGERS_LOCAUX","113_TRIGGERS_MICROSCENES","123_TRIGGERS_RESSOURCES"],
    "narration":["09_NARRATION","81_PENSEES","114_NARRATION_MICROSCENES","124_NARRATION_RESSOURCES","134_NARRATION_FAUNE"],
    "personality":["91_AXES"],
    "obsessions":["92_OBSESSIONS","153_OBSESSIONS_RARES"],
    "memories":["93_SOUVENIRS","154_SOUVENIRS_RENCONTRES"],
    "fauna":["131_TEMPLATES_FAUNE"],
    "npcs":["141_TEMPLATES_PNJ"],
}

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--workbook", required=True)
    ap.add_argument("--output", default="compiled")
    args=ap.parse_args()
    source=Path(args.workbook)
    outdir=Path(args.output); outdir.mkdir(parents=True,exist_ok=True)
    wb=load_workbook(source,data_only=False,read_only=False)
    manifest={"cumVersion":"C2.0","schemaVersion":1,"compiledAt":datetime.now(timezone.utc).isoformat(),"sourceWorkbook":source.name,"datasets":{}}
    for dataset,sheets in DATASETS.items():
        rows=[]
        for name in sheets:
            if name in wb.sheetnames:
                rows.extend(table_rows(wb[name]))
        path=outdir/f"mission-{dataset.replace('_','-')}.json"
        path.write_text(json.dumps(rows,ensure_ascii=False,indent=2),encoding="utf-8")
        manifest["datasets"][dataset]={"rows":len(rows),"file":path.name}
    manifest["sourceSha256"]=hashlib.sha256(source.read_bytes()).hexdigest()
    (outdir/"manifest.json").write_text(json.dumps(manifest,ensure_ascii=False,indent=2),encoding="utf-8")
    print(outdir)

if __name__=="__main__":
    main()
